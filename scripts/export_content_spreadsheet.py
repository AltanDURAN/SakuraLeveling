"""Génère un classeur XLSX consolidant TOUT le contenu du jeu.

Pensé pour l'équilibrage par toi-même : feuilles distinctes, filtres sur
chaque colonne (clique la flèche d'en-tête), et un **score de puissance
calculé en direct** par équipement (formule Excel — il se recalcule quand tu
édites une stat). Tu modifies, tu me renvoies le fichier, je réapplique via
`scripts/import_content_spreadsheet.py`.

Feuilles :
    ⚙️ Calibrage   — constantes du score de puissance (NE PAS supprimer)
    🎽 Équipements — TOUS les équipements en une feuille filtrable + power live
    👹 Mobs        — power score + rang calculés en direct (formules)
    🧪 Consommables, 📦 Ressources, 🛠️ Recettes, 🏪 Shop,
    🧬 Classes, 🌳 Skill Tree, 🏷️ Titres, 🌸 Panoplies, 🐉 World Bosses

XLSX = importable direct dans Google Sheets (Fichier > Importer).
    .venv/bin/python scripts/export_content_spreadsheet.py
Sortie : docs/sakura_content.xlsx
"""

from __future__ import annotations

import json
from datetime import UTC, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.domain.services.power_score_service import (
    _DEF_EFFECTIVE_HITS,
    _RANK_MAX,
    _RANK_THRESHOLDS,
    _SCALE,
)

ROOT = Path(__file__).resolve().parent.parent
CONTENT_DIR = ROOT / "app" / "infrastructure" / "content"
OUTPUT = ROOT / "docs" / "sakura_content.xlsx"

# ---- Palette Sakura ----
HEADER_FILL = PatternFill("solid", fgColor="AD1457")     # rose profond
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
ZEBRA_FILL = PatternFill("solid", fgColor="FCE4EC")       # rose pâle
LOCK_FILL = PatternFill("solid", fgColor="ECEFF1")        # gris : colonnes auto
TITLE_FONT = Font(bold=True, size=15, color="AD1457")
WRAP = Alignment(wrap_text=True, vertical="top")

CALIB = "⚙️ Calibrage"          # nom de la feuille de constantes
CALIB_REF = f"'{CALIB}'!"        # préfixe pour les formules

# Catégories d'équipement → libellé FR lisible (pour la colonne Type / filtre).
EQUIPMENT_TYPES = {
    "weapon": "Arme", "shield": "Bouclier", "helmet": "Casque",
    "chest": "Plastron", "legs": "Jambières", "boots": "Bottes",
    "necklace": "Collier", "bracelet": "Bracelet", "ring": "Bague",
    "belt": "Ceinture", "cape": "Cape", "earring": "Boucle d'oreille",
}


def _load(name: str):
    return json.loads((CONTENT_DIR / name).read_text(encoding="utf-8"))


# ---------------------------------------------------------------- helpers ----
def _autosize(ws, max_width: int = 48) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        longest = 10
        for c in col:
            if c.value is None:
                continue
            s = str(c.value)
            if s.startswith("="):       # formule : largeur non représentative
                s = "00000.0"
            longest = max(longest, max(len(line) for line in s.split("\n")))
        ws.column_dimensions[letter].width = min(max(longest + 2, 10), max_width)


def _style_header(ws, row: int = 1) -> None:
    ws.row_dimensions[row].height = 22
    for cell in ws[row]:
        if cell.value is None:
            continue
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _zebra(ws, header_row: int = 1) -> None:
    for idx, row in enumerate(ws.iter_rows(min_row=header_row + 1), start=header_row + 1):
        if idx % 2 == 0:
            for cell in row:
                if cell.fill.fgColor.rgb in (None, "00000000"):
                    cell.fill = ZEBRA_FILL


def _wrap_multiline(ws) -> None:
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, str) and "\n" in cell.value:
                cell.alignment = WRAP


def _finalize(ws, freeze: str = "A2") -> None:
    """Styling commun : header, zebra, autofilter, freeze, autosize."""
    _style_header(ws)
    _wrap_multiline(ws)
    _zebra(ws)
    _autosize(ws)
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = freeze


# ----------------------------------------------------------- calibrage --------
def write_calibrage_sheet(wb: Workbook) -> None:
    """Constantes du score de puissance + table des rangs (pour VLOOKUP).

    Les formules « Power » des autres feuilles pointent ici → édite une base
    et tout se recalcule. Miroir EXACT de PowerScoreService.
    """
    ws = wb.create_sheet(CALIB)
    ws["A1"] = "⚙️ Calibrage du score de puissance — NE PAS supprimer (formules liées)"
    ws["A1"].font = TITLE_FONT

    rows = [
        ("Base ATK", 10), ("Base DEF", 5), ("Base PV", 100),
        ("Base Crit %", 5), ("Base Crit dmg (100 = neutre)", 150),
        ("Base Esquive %", 0), ("Base Vitesse", 5),
        ("K_DEF (1 DEF = N PV effectifs)", _DEF_EFFECTIVE_HITS),
        ("SCALE (diviseur)", _SCALE),
    ]
    for i, (label, val) in enumerate(rows, start=2):   # B2..B10
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=val)
    # B11 : power du build de base seul (référence pour le power MARGINAL).
    ws["A11"] = "Power base (auto)"
    ws["B11"] = (
        "=$B$2*(1+($B$5/100)*MAX(0,$B$6-100)/100)*(1+$B$8/100)"
        "*(($B$4+$B$3*$B$9)/MAX(0.01,1-$B$7/100))/$B$10"
    )
    for r in range(2, 12):
        ws.cell(row=r, column=1).font = Font(bold=True)

    # Table des rangs : (score ≥ borne_basse) → rang, pour VLOOKUP approx.
    ws["D1"] = "Score ≥"
    ws["E1"] = "Rang"
    table = [(0, _RANK_THRESHOLDS[0][1])]
    for i in range(1, len(_RANK_THRESHOLDS)):
        table.append((_RANK_THRESHOLDS[i - 1][0], _RANK_THRESHOLDS[i][1]))
    table.append((_RANK_THRESHOLDS[-1][0], _RANK_MAX))
    for i, (lb, rank) in enumerate(table, start=2):
        ws.cell(row=i, column=4, value=lb)
        ws.cell(row=i, column=5, value=rank)
    for cell in (ws["D1"], ws["E1"]):
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 8
    global _RANK_TABLE_LAST
    _RANK_TABLE_LAST = len(table) + 1   # dernière ligne de la table (offset header)


_RANK_TABLE_LAST = 27


def _equip_power_formula(r: int) -> str:
    """Power MARGINAL d'un équipement (= power(base+item) − power(base))."""
    c = CALIB_REF
    atk = f"({c}$B$2+H{r})"
    deff = f"({c}$B$3+I{r})"
    hp = f"({c}$B$4+J{r})"
    cc = f"({c}$B$5+K{r})"
    cd = f"({c}$B$6+L{r})"
    dg = f"({c}$B$7+M{r})"
    sp = f"({c}$B$8+N{r})"
    off = f"{atk}*(1+({cc}/100)*MAX(0,{cd}-100)/100)*(1+{sp}/100)"
    ehp = f"({hp}+{deff}*{c}$B$9)/MAX(0.01,1-{dg}/100)"
    return f"=ROUND(({off})*({ehp})/{c}$B$10-{c}$B$11,1)"


def _mob_power_formula(r: int) -> str:
    """Power ABSOLU d'un mob (colonnes D..J de la feuille Mobs)."""
    c = CALIB_REF
    off = f"E{r}*(1+(H{r}/100)*MAX(0,I{r}-100)/100)*(1+G{r}/100)"
    ehp = f"(D{r}+F{r}*{c}$B$9)/MAX(0.01,1-J{r}/100)"
    return f"=ROUND(({off})*({ehp})/{c}$B$10,0)"


def _mob_rank_formula(r: int) -> str:
    return (
        f"=IFERROR(VLOOKUP(O{r},{CALIB_REF}$D$2:$E${_RANK_TABLE_LAST},2,TRUE),\"?\")"
    )


def _power_color_scale(ws, col_letter: str, last_row: int) -> None:
    if last_row < 2:
        return
    ws.conditional_formatting.add(
        f"{col_letter}2:{col_letter}{last_row}",
        ColorScaleRule(
            start_type="min", start_color="63BE7B",          # vert (faible)
            mid_type="percentile", mid_value=50, mid_color="FFEB84",
            end_type="max", end_color="F8696B",              # rouge (fort)
        ),
    )


# --------------------------------------------------------- équipements --------
def write_equipment_sheet(wb: Workbook) -> None:
    items = _load("items.json")
    equip = [i for i in items if i.get("category") in EQUIPMENT_TYPES]
    # tri : type (ordre canonique) puis panoplie puis nom
    type_order = list(EQUIPMENT_TYPES)
    equip.sort(key=lambda it: (
        type_order.index(it["category"]),
        it.get("family") or "~", it.get("name", ""),
    ))

    ws = wb.create_sheet("🎽 Équipements")
    headers = [
        "Code", "Nom", "Type", "Slot", "Panoplie", "Rareté", "2-mains",
        "Atk", "Def", "PV", "Crit %", "Crit dmg", "Esquive", "Vitesse", "Regen",
        "Power", "Prix achat", "Description",
    ]
    ws.append(headers)

    for it in equip:
        b = it.get("stat_bonuses") or {}
        ws.append([
            it["code"], it["name"],
            EQUIPMENT_TYPES.get(it["category"], it["category"]),
            it.get("equipment_slot") or "—",
            it.get("family") or "",
            it.get("rarity", "—"),
            "oui" if it.get("requires_two_hands") else "non",
            b.get("attack", 0), b.get("defense", 0), b.get("max_hp", 0),
            b.get("crit_chance", 0), b.get("crit_damage", 0),
            b.get("dodge", 0), b.get("speed", 0), b.get("hp_regeneration", 0),
            None,  # Power : formule injectée ci-dessous
            it.get("buy_price") or 0,
            it.get("description", ""),
        ])

    last = ws.max_row
    for r in range(2, last + 1):
        ws.cell(row=r, column=16, value=_equip_power_formula(r))   # P = Power
        ws.cell(row=r, column=16).number_format = "0.0"
        ws.cell(row=r, column=16).fill = LOCK_FILL                 # auto-calc
        ws.cell(row=r, column=16).font = Font(bold=True)

    _finalize(ws, freeze="C2")
    _power_color_scale(ws, "P", last)
    ws.column_dimensions["P"].width = 9


# ------------------------------------------------------------- mobs -----------
def write_mobs_sheet(wb: Workbook) -> None:
    mobs = _load("mobs.json")
    ws = wb.create_sheet("👹 Mobs")
    ws.append([
        "Code", "Nom", "Famille",
        "PV max", "Atk", "Def", "Vit", "Crit %", "Crit dmg", "Esquive", "Regen",
        "XP récomp.", "Or récomp.", "Poids spawn",
        "Power", "Rang", "Drops", "Image", "Description",
    ])
    for m in mobs:
        drops = "\n".join(
            f"{d['item_code']} ×{d.get('min_quantity',1)}-{d.get('max_quantity',1)} "
            f"({int(float(d['drop_rate'])*100)}%)"
            for d in (m.get("loot_table") or [])
        )
        ws.append([
            m["code"], m["name"], m.get("family", ""),
            m["max_hp"], m["attack"], m["defense"], m["speed"],
            m["crit_chance"], m["crit_damage"], m["dodge"],
            m.get("hp_regeneration", 0),
            m.get("xp_reward", 0), m.get("gold_reward", 0), m.get("spawn_weight", 1),
            None, None,            # Power / Rang : formules
            drops, m.get("image_name", ""), m.get("description", ""),
        ])
    last = ws.max_row
    for r in range(2, last + 1):
        ws.cell(row=r, column=15, value=_mob_power_formula(r))   # O = Power
        ws.cell(row=r, column=16, value=_mob_rank_formula(r))    # P = Rang
        ws.cell(row=r, column=15).fill = LOCK_FILL
        ws.cell(row=r, column=16).fill = LOCK_FILL
        ws.cell(row=r, column=15).font = Font(bold=True)
        ws.cell(row=r, column=16).font = Font(bold=True)
    _finalize(ws, freeze="C2")
    _power_color_scale(ws, "O", last)


# --------------------------------------------------- consommables / ressources
def write_consumables_sheet(wb: Workbook) -> None:
    rows = [i for i in _load("items.json") if i.get("category") == "consumable"]
    if not rows:
        return
    ws = wb.create_sheet("🧪 Consommables")
    ws.append(["Code", "Nom", "Rareté", "Effet", "Valeur",
               "Prix vente", "Prix achat", "Description"])
    for it in rows:
        b = it.get("stat_bonuses") or {}
        ws.append([it["code"], it["name"], it.get("rarity", "—"),
                   b.get("effect", ""), b.get("value", ""),
                   it.get("sell_price") or 0, it.get("buy_price") or 0,
                   it.get("description", "")])
    _finalize(ws, freeze="B2")


def write_resources_sheet(wb: Workbook) -> None:
    rows = [i for i in _load("items.json") if i.get("category") == "resource"]
    if not rows:
        return
    ws = wb.create_sheet("📦 Ressources")
    ws.append(["Code", "Nom", "Rareté", "Stackable", "Stack max",
               "Prix vente", "Prix achat", "Description"])
    for it in rows:
        ws.append([it["code"], it["name"], it.get("rarity", "—"),
                   "oui" if it.get("stackable") else "non",
                   it.get("max_stack") or "∞",
                   it.get("sell_price") or 0, it.get("buy_price") or 0,
                   it.get("description", "")])
    _finalize(ws, freeze="B2")


# ----------------------------------------------------------- autres ----------
def write_crafts_sheet(wb: Workbook) -> None:
    crafts = _load("crafts.json")
    items = {i["code"]: i for i in _load("items.json")}
    ws = wb.create_sheet("🛠️ Recettes")
    ws.append(["Code recette", "Nom", "Résultat (item)", "Catégorie résultat",
               "Slot résultat", "Quantité produite", "Ingrédients"])
    for r in crafts:
        result = items.get(r["result_item_code"])
        ws.append([
            r["code"], r["name"], r["result_item_code"],
            (result or {}).get("category", "?"),
            (result or {}).get("equipment_slot") or "—",
            r.get("result_quantity", 1),
            "\n".join(f"{ing['item_code']} ×{ing['quantity']}"
                      for ing in r.get("ingredients", [])),
        ])
    _finalize(ws, freeze="B2")


def write_shop_sheet(wb: Workbook) -> None:
    try:
        shop = _load("shop_items.json")
    except FileNotFoundError:
        return
    items = {i["code"]: i for i in _load("items.json")}
    ws = wb.create_sheet("🏪 Shop")
    ws.append(["Code item", "Nom", "Catégorie", "Prix achat",
               "Prix vente max", "Prix vente min", "Seuil stock", "Activé"])
    for e in shop:
        it = items.get(e["item_code"]) or {}
        ws.append([e["item_code"], it.get("name", "?"), it.get("category", "?"),
                   e.get("buy_price", 0), e.get("max_sell_price", 0),
                   e.get("min_sell_price", 0), e.get("stock_threshold", 0),
                   "oui" if e.get("enabled", True) else "non"])
    _finalize(ws, freeze="B2")


def write_classes_sheet(wb: Workbook) -> None:
    classes = _load("classes.json")
    ws = wb.create_sheet("🧬 Classes")
    ws.append(["Code", "Nom", "Atk", "Def", "PV", "Crit %", "Crit dmg",
               "Esquive", "Vitesse", "Regen", "Déblocage", "Description"])
    for c in classes:
        b = c.get("stat_bonuses") or {}
        reqs = []
        for r in c.get("unlock_requirements", []):
            t = r.get("type", "?")
            if t == "profession_level":
                reqs.append(f"métier {r.get('profession_code')} niv. {r.get('level')}")
            elif t == "item":
                reqs.append(f"item {r.get('item_code')}")
            else:
                reqs.append(json.dumps(r, ensure_ascii=False))
        ws.append([c["code"], c["name"],
                   b.get("attack", 0), b.get("defense", 0), b.get("max_hp", 0),
                   b.get("crit_chance", 0), b.get("crit_damage", 0),
                   b.get("dodge", 0), b.get("speed", 0), b.get("hp_regeneration", 0),
                   "\n".join(reqs) or "aucune", c.get("description", "")])
    _finalize(ws, freeze="B2")


def write_skill_tree_sheet(wb: Workbook) -> None:
    try:
        tree = _load("skill_tree.json")
    except FileNotFoundError:
        return
    ws = wb.create_sheet("🌳 Skill Tree")
    ws.append(["Code", "Nom", "Niveau max", "Coûts/niveau", "Effets",
               "Prérequis", "Position", "Icône", "Description"])
    for code, node in tree.get("skills", {}).items():
        effects = "\n".join(
            f"{e.get('type','?')} → [{','.join(str(v) for v in e.get('values', []))}]"
            for e in node.get("effects", [])
        )
        pos = node.get("position", {})
        ws.append([code, node.get("name", ""), node.get("max_level", 0),
                   ",".join(str(c) for c in node.get("costs", [])),
                   effects or "—",
                   ", ".join(node.get("prerequisites", [])) or "racine",
                   f"{pos.get('x', 0)}, {pos.get('y', 0)}",
                   node.get("icon", ""), node.get("description", "")])
    _finalize(ws, freeze="B2")


def write_titles_sheet(wb: Workbook) -> None:
    try:
        titles = _load("titles.json")
    except FileNotFoundError:
        return
    ws = wb.create_sheet("🏷️ Titres")
    ws.append(["Code", "Nom", "Icône", "Exclusif", "Condition",
               "Cible / Valeur", "Effets", "Description"])
    for t in titles:
        target, value = t.get("condition_target", ""), t.get("condition_value", 0)
        cond_target = (f"{target} / {value}" if target and value
                       else (str(value) if value else (target or "—")))
        effects = "\n".join(
            f"{e.get('type','?')} → "
            + (f"target={e['target']}, " if e.get("target") else "")
            + f"value={e.get('value', 0)}"
            for e in (t.get("effects") or [])
        )
        ws.append([t["code"], t["name"], t.get("icon", ""),
                   "oui" if t.get("exclusive") else "non",
                   t.get("condition_type", "—"), cond_target,
                   effects or "—", t.get("description", "")])
    _finalize(ws, freeze="B2")


def write_sets_sheet(wb: Workbook) -> None:
    try:
        sets_def = _load("sets.json")
    except FileNotFoundError:
        return
    items = _load("items.json")
    counts: dict[str, int] = {}
    for it in items:
        fam = (it.get("family") or "").strip()
        if fam and it.get("equipment_slot"):
            counts[fam] = counts.get(fam, 0) + 1
    ws = wb.create_sheet("🌸 Panoplies")
    ws.append(["Code famille", "Nom", "Icône", "Items existants", "Description",
               "Palier 2", "Palier 4", "Palier 8", "Palier 12"])
    for code, sd in sets_def.items():
        cells = ["—", "—", "—", "—"]
        for tier in sd.get("tiers", []):
            mp = int(tier.get("min_pieces", 0))
            label = f"{tier.get('type', '?')} +{tier.get('value', 0)}"
            cells[{2: 0, 4: 1, 8: 2, 12: 3}.get(mp, 0)] = label
        ws.append([code, sd.get("name", code), sd.get("icon", ""),
                   counts.get(code, 0), sd.get("description", ""), *cells])
    _finalize(ws, freeze="B2")


def write_world_bosses_sheet(wb: Workbook) -> None:
    try:
        bosses = _load("boss_definitions.json")
    except FileNotFoundError:
        return
    ws = wb.create_sheet("🐉 World Bosses")
    ws.append(["Code", "Nom", "Famille", "PV max", "Atk", "Def", "Vit",
               "Crit %", "Crit dmg", "Esquive", "Modifiers", "Poids spawn", "Lore"])
    for b in bosses:
        mods = "\n".join(f"{k} = {v}" for k, v in (b.get("modifiers") or {}).items())
        ws.append([b["code"], b["name"], b.get("family", ""),
                   b["max_hp"], b["attack"], b["defense"], b["speed"],
                   b["crit_chance"], b["crit_damage"], b["dodge"],
                   mods or "—", b.get("spawn_weight", 1),
                   b.get("description", "")[:300]])
    _finalize(ws, freeze="C2")


def write_summary_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("📋 Sommaire", 0)
    ws["A1"] = "🌸 SakuraLeveling V2 — Contenu du jeu"
    ws["A1"].font = Font(bold=True, size=18, color="AD1457")
    ws["A2"] = f"Généré le {datetime.now(UTC).strftime('%Y-%m-%d %H:%M UTC')}"
    ws["A2"].font = Font(italic=True, color="888888")

    ws.append([])
    ws.append(["Feuille", "Contenu"])
    _style_header(ws, row=4)
    feuilles = [
        ("🎽 Équipements", "TOUS les équipements en 1 feuille. Filtre par Type "
         "(casque…) ou Panoplie via les flèches d'en-tête. Power calculé en direct."),
        ("👹 Mobs", "Monstres : stats, drops, Power & Rang calculés en direct."),
        ("🧪 Consommables / 📦 Ressources", "Potions et ressources de craft."),
        ("🛠️ Recettes", "Recettes /craft et /forge."),
        ("🏪 Shop", "Prix d'achat / vente du shop."),
        ("🧬 Classes", "Classes joueur + bonus + déblocage."),
        ("🌳 Skill Tree", "Tous les nœuds de l'arbre."),
        ("🏷️ Titres", "Titres, conditions et effets."),
        ("🌸 Panoplies", "Familles de set + paliers de bonus."),
        ("🐉 World Bosses", "Bosses + modifiers."),
        ("⚙️ Calibrage", "Constantes du score de puissance. NE PAS supprimer."),
    ]
    for name, desc in feuilles:
        ws.append([name, desc])

    ws.append([])
    ws.append(["🛠️ Comment équilibrer (round-trip)"])
    ws[f"A{ws.max_row}"].font = Font(bold=True, size=13, color="AD1457")
    notes = [
        "1. Filtre : clique la flèche d'un en-tête (ex : colonne Type = « Casque », "
        "ou Panoplie = « gobelin »).",
        "2. Édite les colonnes de stats (Atk, Def, PV, Crit %, Crit dmg, Esquive, "
        "Vitesse, Regen), le Prix, la Rareté, le Slot, la Panoplie, le Nom…",
        "3. La colonne « Power » (fond gris) est AUTO-CALCULÉE — ne l'édite pas, "
        "elle se met à jour toute seule quand tu changes une stat.",
        "4. Tu peux AJOUTER une ligne d'équipement : remplis au moins Code + Type. "
        "Tu peux en supprimer aussi.",
        "5. Renvoie-moi le fichier .xlsx modifié → je réapplique via "
        "scripts/import_content_spreadsheet.py (puis seed + déploiement).",
        "",
        "Conventions stats : crit_chance & esquive en 0..100 ; crit_damage 100 = "
        "neutre, 150 = ×1,5 ; défense soustractive (1 DEF ≈ 25 PV effectifs).",
        "Power d'un équipement = puissance MARGINALE qu'il ajoute à un perso de base "
        "(power(base+item) − power(base)).",
    ]
    for n in notes:
        ws.append([n])
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 90
    ws.freeze_panes = "A5"


def main() -> None:
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    write_calibrage_sheet(wb)      # d'abord : définit _RANK_TABLE_LAST
    write_equipment_sheet(wb)
    write_mobs_sheet(wb)
    write_consumables_sheet(wb)
    write_resources_sheet(wb)
    write_crafts_sheet(wb)
    write_shop_sheet(wb)
    write_classes_sheet(wb)
    write_skill_tree_sheet(wb)
    write_titles_sheet(wb)
    write_sets_sheet(wb)
    write_world_bosses_sheet(wb)
    write_summary_sheet(wb)        # index 0

    # Ordre d'affichage : Sommaire, Équipements, Mobs… puis Calibrage en dernier.
    wb.move_sheet(CALIB, offset=len(wb.sheetnames))

    wb.save(OUTPUT)
    print(f"✅ Tableur généré : {OUTPUT}")
    print(f"   {len(wb.sheetnames)} feuilles : {', '.join(wb.sheetnames)}")


if __name__ == "__main__":
    main()
