"""Réapplique un classeur XLSX édité vers les JSON de contenu du jeu.

Pendant de `export_content_spreadsheet.py` : tu édites les feuilles à stats
PLATES (Équipements, Mobs, World Bosses, Classes, Shop), tu me renvoies le
.xlsx, je relance ce script qui réécrit items.json / mobs.json / etc.

Round-trip par CODE (colonne Code) : ajout, modification et suppression de
lignes sont tous gérés. Les colonnes calculées (Power, Rang) sont ignorées.
Les champs non présents dans le tableur (icon, loot_table, image…) sont
PRÉSERVÉS. Les feuilles structurées (Skill Tree, Titres, Panoplies, Recettes)
ne sont PAS ré-importées ici — dis-moi les changements à la main.

    .venv/bin/python scripts/import_content_spreadsheet.py [chemin.xlsx]
        --dry-run   n'écrit rien, affiche seulement le diff
        --no-delete ne supprime aucune entrée absente du tableur
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
CONTENT_DIR = ROOT / "app" / "infrastructure" / "content"
DEFAULT_XLSX = ROOT / "docs" / "sakura_content.xlsx"

EQUIPMENT_TYPES = {
    "weapon": "Arme", "shield": "Bouclier", "helmet": "Casque",
    "chest": "Plastron", "legs": "Jambières", "boots": "Bottes",
    "necklace": "Collier", "bracelet": "Bracelet", "ring": "Bague",
    "belt": "Ceinture", "cape": "Cape", "earring": "Boucle d'oreille",
}
_TYPE_TO_CAT = {v: k for k, v in EQUIPMENT_TYPES.items()}

# stat de la feuille → clé stat_bonuses
_STAT_COLS = {
    "Atk": "attack", "Def": "defense", "PV": "max_hp", "Crit %": "crit_chance",
    "Crit dmg": "crit_damage", "Esquive": "dodge", "Vitesse": "speed",
    "Regen": "hp_regeneration",
}


def _load(name):
    return json.loads((CONTENT_DIR / name).read_text(encoding="utf-8"))


def _save(name, data):
    (CONTENT_DIR / name).write_text(
        json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def _sheet_rows(wb, name):
    """Yield des dicts {header: value} pour une feuille (None si absente)."""
    if name not in wb.sheetnames:
        return None
    ws = wb[name]
    headers = [c.value for c in ws[1]]
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        out.append({h: v for h, v in zip(headers, row) if h is not None})
    return out


def _i(v, default=0):
    if v in (None, "", "—", "∞"):
        return default
    try:
        return int(round(float(v)))
    except (TypeError, ValueError):
        return default


def _s(v, default=""):
    return default if v is None else str(v).strip()


def _bool(v):
    return _s(v).lower() in ("oui", "true", "1", "vrai", "yes")


def _stats_from_row(row):
    out = {}
    for col, key in _STAT_COLS.items():
        val = _i(row.get(col))
        if val:
            out[key] = val
    return out


# --------------------------------------------------------- équipements --------
def apply_equipment(wb, delete: bool, log: list[str]) -> bool:
    rows = _sheet_rows(wb, "🎽 Équipements")
    if rows is None:
        return False
    items = _load("items.json")
    by_code = {it["code"]: it for it in items}
    seen = set()
    changed = added = removed = 0

    for row in rows:
        code = _s(row.get("Code"))
        if not code:
            continue
        seen.add(code)
        cat = _TYPE_TO_CAT.get(_s(row.get("Type")))
        slot = _s(row.get("Slot")) or None
        if slot == "—":
            slot = None
        patch = {
            "name": _s(row.get("Nom")) or code,
            "rarity": _s(row.get("Rareté")) or "common",
            "equipment_slot": slot,
            "family": _s(row.get("Panoplie")),
            "requires_two_hands": _bool(row.get("2-mains")),
            "stat_bonuses": _stats_from_row(row),
            "buy_price": _i(row.get("Prix achat")) or None,
            "description": _s(row.get("Description")),
        }
        if code in by_code:
            it = by_code[code]
            before = json.dumps(it, sort_keys=True, ensure_ascii=False)
            it.update(patch)
            if cat:
                it["category"] = cat
            if json.dumps(it, sort_keys=True, ensure_ascii=False) != before:
                changed += 1
                log.append(f"  ~ {code}")
        else:
            if not cat:
                log.append(f"  ! ligne ignorée (Type inconnu) : {code}")
                continue
            new = {
                "code": code, "name": patch["name"],
                "description": patch["description"], "category": cat,
                "rarity": patch["rarity"], "stackable": False, "max_stack": None,
                "sell_price": 0, "buy_price": patch["buy_price"], "icon": None,
                "stat_bonuses": patch["stat_bonuses"],
                "equipment_slot": patch["equipment_slot"],
                "requires_two_hands": patch["requires_two_hands"],
                "family": patch["family"],
            }
            items.append(new)
            by_code[code] = new
            added += 1
            log.append(f"  + {code} ({cat})")

    if delete:
        for it in list(items):
            if it.get("category") in EQUIPMENT_TYPES and it["code"] not in seen:
                items.remove(it)
                removed += 1
                log.append(f"  - {it['code']}")

    log.insert(0, f"🎽 Équipements : {changed} modifiés, {added} ajoutés, "
                  f"{removed} supprimés")
    if changed or added or removed:
        _save_pending["items.json"] = items
        return True
    return False


# ------------------------------------------------- mobs / bosses / classes ----
def _apply_stat_sheet(wb, sheet, json_name, fields, key_col, log):
    """Générique pour une feuille à stats plates indexée par code."""
    rows = _sheet_rows(wb, sheet)
    if rows is None:
        return False
    data = _load(json_name)
    by_code = {d["code"]: d for d in data}
    changed = 0
    for row in rows:
        code = _s(row.get(key_col))
        entry = by_code.get(code)
        if not entry:
            continue
        before = json.dumps(entry, sort_keys=True, ensure_ascii=False)
        for col, key in fields.items():
            if col in row:
                entry[key] = _i(row.get(col))
        if json_name == "mobs.json":
            entry["current_hp"] = entry.get("max_hp", entry.get("current_hp", 0))
        if json.dumps(entry, sort_keys=True, ensure_ascii=False) != before:
            changed += 1
            log.append(f"  ~ {code}")
    log.insert(0, f"{sheet} : {changed} modifiés")
    if changed:
        _save_pending[json_name] = data
        return True
    return False


_MOB_FIELDS = {
    "PV max": "max_hp", "Atk": "attack", "Def": "defense", "Vit": "speed",
    "Crit %": "crit_chance", "Crit dmg": "crit_damage", "Esquive": "dodge",
    "Regen": "hp_regeneration", "XP récomp.": "xp_reward",
    "Or récomp.": "gold_reward", "Poids spawn": "spawn_weight",
}
_BOSS_FIELDS = {
    "PV max": "max_hp", "Atk": "attack", "Def": "defense", "Vit": "speed",
    "Crit %": "crit_chance", "Crit dmg": "crit_damage", "Esquive": "dodge",
    "Poids spawn": "spawn_weight",
}


def apply_classes(wb, log):
    rows = _sheet_rows(wb, "🧬 Classes")
    if rows is None:
        return False
    data = _load("classes.json")
    by_code = {c["code"]: c for c in data}
    fields = {"Atk": "attack", "Def": "defense", "PV": "max_hp",
              "Crit %": "crit_chance", "Crit dmg": "crit_damage",
              "Esquive": "dodge", "Vitesse": "speed", "Regen": "hp_regeneration"}
    changed = 0
    for row in rows:
        c = by_code.get(_s(row.get("Code")))
        if not c:
            continue
        before = json.dumps(c, sort_keys=True, ensure_ascii=False)
        bonuses = {k: _i(row.get(col)) for col, k in fields.items() if _i(row.get(col))}
        c["stat_bonuses"] = bonuses
        if json.dumps(c, sort_keys=True, ensure_ascii=False) != before:
            changed += 1
            log.append(f"  ~ {c['code']}")
    log.insert(0, f"🧬 Classes : {changed} modifiés")
    if changed:
        _save_pending["classes.json"] = data
        return True
    return False


def apply_shop(wb, log):
    rows = _sheet_rows(wb, "🏪 Shop")
    if rows is None:
        return False
    data = _load("shop_items.json")
    by_code = {s["item_code"]: s for s in data}
    changed = 0
    for row in rows:
        s = by_code.get(_s(row.get("Code item")))
        if not s:
            continue
        before = json.dumps(s, sort_keys=True, ensure_ascii=False)
        s["buy_price"] = _i(row.get("Prix achat"))
        s["max_sell_price"] = _i(row.get("Prix vente max"))
        s["min_sell_price"] = _i(row.get("Prix vente min"))
        s["stock_threshold"] = _i(row.get("Seuil stock"))
        s["enabled"] = _bool(row.get("Activé"))
        if json.dumps(s, sort_keys=True, ensure_ascii=False) != before:
            changed += 1
            log.append(f"  ~ {s['item_code']}")
    log.insert(0, f"🏪 Shop : {changed} modifiés")
    if changed:
        _save_pending["shop_items.json"] = data
        return True
    return False


_save_pending: dict[str, object] = {}


def main() -> None:
    args = sys.argv[1:]
    dry = "--dry-run" in args
    delete = "--no-delete" not in args
    paths = [a for a in args if not a.startswith("--")]
    xlsx = Path(paths[0]) if paths else DEFAULT_XLSX
    if not xlsx.exists():
        sys.exit(f"❌ Introuvable : {xlsx}")

    wb = load_workbook(xlsx, data_only=True)
    blocks: list[list[str]] = []

    # Chaque importeur écrit dans son propre log puis on l'agrège pour l'affichage.
    runners = [
        lambda lg: apply_equipment(wb, delete, lg),
        lambda lg: _apply_stat_sheet(wb, "👹 Mobs", "mobs.json", _MOB_FIELDS, "Code", lg),
        lambda lg: _apply_stat_sheet(wb, "🐉 World Bosses", "boss_definitions.json", _BOSS_FIELDS, "Code", lg),
        lambda lg: apply_classes(wb, lg),
        lambda lg: apply_shop(wb, lg),
    ]
    for runner in runners:
        lg: list[str] = []
        runner(lg)
        if lg:
            blocks.append(lg)

    print(f"📄 Source : {xlsx}")
    for block in blocks:
        for line in block:
            print(line)

    if dry:
        print("\n🔍 --dry-run : aucun fichier écrit.")
        return
    if not _save_pending:
        print("\nℹ️ Rien à écrire (aucun changement détecté).")
        return
    for name, data in _save_pending.items():
        _save(name, data)
    print(f"\n✅ Écrit : {', '.join(_save_pending)}")
    print("Pense à : seed_content + déploiement (et re-export du tableur).")


if __name__ == "__main__":
    main()
